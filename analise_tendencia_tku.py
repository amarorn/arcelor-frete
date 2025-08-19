#!/usr/bin/env python3
"""
Análise de Tendência de Preços TKU por Micro-região
Gera relatórios de tendência para todas as micro-regiões
"""

import os
import time
import logging
from datetime import datetime, timedelta
import pandas as pd
import numpy as np
from typing import Dict, List, Tuple, Optional

# Configurar logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class AnalisadorTendenciaTKU:
    """Analisa tendências de preços TKU por micro-região ao longo do tempo"""
    
    def __init__(self, data_path: str = "sample_data.xlsx"):
        self.data_path = data_path
        self.df = None
        self.microregioes_disponiveis = []
        
    def carregar_dados(self) -> bool:
        """Carrega dados do Excel"""
        try:
            # Tentar múltiplos caminhos para o arquivo
            possible_paths = [
                self.data_path,
                f"../{self.data_path}",
                f"../../{self.data_path}",
                "sample_data.xlsx",
                "../sample_data.xlsx",
                "../../sample_data.xlsx"
            ]
            
            file_found = False
            for path in possible_paths:
                if os.path.exists(path):
                    logger.info(f"Arquivo encontrado em: {path}")
                    self.data_path = path
                    file_found = True
                    break
            
            if not file_found:
                logger.error("Arquivo sample_data.xlsx não encontrado!")
                return False
            
            # Carregar dados
            self.df = pd.read_excel(self.data_path)
            logger.info(f"Dados carregados: {len(self.df)} linhas")
            
            # Preparar dados
            self._preparar_dados()
            return True
            
        except Exception as e:
            logger.error(f"Erro ao carregar dados: {str(e)}")
            return False
    
    def _preparar_dados(self):
        """Prepara e limpa os dados"""
        # Mapeamento de colunas
        column_mapping = {
            '00.dt_doc_faturamento': 'data_faturamento',
            '00.nm_centro_origem_unidade': 'centro_origem',
            'dc_centro_unidade_descricao': 'descricao_centro',
            '00.nm_modal': 'modal',
            'nm_tipo_rodovia': 'tipo_rodovia',
            'nm_veiculo': 'tipo_veiculo',
            '01.Rota_UFOrigem_UFDestino': 'rota_uf',
            '01.Rota_MesoOrigem_MesoDestino': 'rota_mesoregiao',
            '01.Rota_MicroOrigem_MicroDestino': 'rota_microregiao',
            '01.Rota_MuniOrigem_MuniDestino': 'rota_municipio',
            'id_transportadora': 'transportadora',
            'Date': 'data',
            'id_fatura': 'fatura',
            'id_transporte': 'transporte',
            '02.01.00 - Volume (ton)': 'volume_ton',
            '02.01.01 - Frete Geral (BRL)': 'frete_brl',
            '02.01.02 - DISTANCIA (KM)': 'distancia_km',
            '02.03.00 - Preço_Frete Geral (BRL) / TON': 'preco_ton',
            '02.03.02 - Preço_Frete Geral (BRL / TON / KM)': 'preco_ton_km'
        }
        
        # Renomear colunas
        self.df = self.df.rename(columns=column_mapping)
        
        # Converter tipos
        self.df['data_faturamento'] = pd.to_datetime(self.df['data_faturamento'], errors='coerce')
        self.df['volume_ton'] = pd.to_numeric(self.df['volume_ton'], errors='coerce')
        self.df['distancia_km'] = pd.to_numeric(self.df['distancia_km'], errors='coerce')
        self.df['preco_ton_km'] = pd.to_numeric(self.df['preco_ton_km'], errors='coerce')
        
        # Filtrar dados válidos
        self.df = self.df[
            (self.df['volume_ton'] > 0) & 
            (self.df['distancia_km'] > 0) & 
            (self.df['preco_ton_km'] > 0) &
            (self.df['data_faturamento'].notna())
        ].copy()
        
        # Extrair micro-região
        self.df['microregiao_origem'] = self.df['centro_origem'].apply(self._extrair_microregiao)
        
        # Criar período mensal para análise temporal
        self.df['ano_mes'] = self.df['data_faturamento'].dt.to_period('M')
        
        # Listar micro-regiões disponíveis
        self.microregioes_disponiveis = sorted(self.df['microregiao_origem'].unique())
        
        logger.info(f"Dados preparados: {len(self.df)} linhas válidas")
        logger.info(f"Micro-regiões identificadas: {len(self.microregioes_disponiveis)}")
        
        # Mostrar estatísticas gerais
        print(f"\n📊 ESTATÍSTICAS GERAIS:")
        print(f"   • Total de rotas: {len(self.df):,}")
        print(f"   • Período: {self.df['ano_mes'].min()} a {self.df['ano_mes'].max()}")
        print(f"   • Preço médio geral: R$ {self.df['preco_ton_km'].mean():.4f}/ton.km")
        print(f"   • Volume total: {self.df['volume_ton'].sum():,.1f} ton")
        print(f"   • Distância média: {self.df['distancia_km'].mean():.1f} km")
    
    def _extrair_microregiao(self, centro_origem: str) -> str:
        """Extrai micro-região do centro de origem"""
        if pd.isna(centro_origem):
            return "UNKNOWN"
        
        centro_str = str(centro_origem).upper()
        
        # Mapeamento de micro-regiões - Minas Gerais
        microregion_mapping = [
            ('JOÃO MONLEVADE', 'JOÃO MONLEVADE'),
            ('USINA MONLEVADE', 'JOÃO MONLEVADE'),
            ('USINA', 'ITABIRA'),  # Mapear USINA para ITABIRA
            ('ITABIRA', 'ITABIRA'),
            ('BELO HORIZONTE', 'BELO HORIZONTE'),
            ('CONTAGEM', 'CONTAGEM'),
            ('SABARÁ', 'SABARÁ'),
            ('SANTA LUZIA', 'SANTA LUZIA'),
            ('NOVA LIMA', 'NOVA LIMA'),
            ('BRUMADINHO', 'BRUMADINHO'),
            ('IBIRITÉ', 'IBIRITÉ'),
            ('BETIM', 'BETIM'),
            ('LAGOA SANTA', 'LAGOA SANTA'),
            ('VESPASIANO', 'VESPASIANO'),
            ('RIBEIRÃO DAS NEVES', 'RIBEIRÃO DAS NEVES'),
            ('CAETÉ', 'CAETÉ'),
            ('SÃO JOSÉ DA LAPA', 'SÃO JOSÉ DA LAPA'),
            ('FLORESTAL', 'FLORESTAL'),
            ('JABOTICATUBAS', 'JABOTICATUBAS'),
            ('MATEUS LEME', 'MATEUS LEME'),
            ('IGARAPÉ', 'IGARAPÉ'),
            ('SÃO JOAQUIM DE BICAS', 'SÃO JOAQUIM DE BICAS'),
            ('SÃO JOSÉ DO GOIABAL', 'SÃO JOSÉ DO GOIABAL'),
            ('MARAVILHAS', 'MARAVILHAS'),
            ('ONÇA DE PITANGUI', 'ONÇA DE PITANGUI'),
            ('PARÁ DE MINAS', 'PARÁ DE MINAS'),
            ('PITANGUI', 'PITANGUI'),
            ('CONCEIÇÃO DO MATO DENTRO', 'CONCEIÇÃO DO MATO DENTRO'),
            ('SANTANA DO PARAÍSO', 'SANTANA DO PARAÍSO'),
            ('CORONEL FABRICIANO', 'CORONEL FABRICIANO'),
            ('IPATINGA', 'IPATINGA'),
            ('TIMÓTEO', 'TIMÓTEO'),
            ('CARATINGA', 'CARATINGA'),
            ('INHAPIM', 'INHAPIM'),
            ('GOVERNADOR VALADARES', 'GOVERNADOR VALADARES'),
            ('TEÓFILO OTONI', 'TEÓFILO OTONI'),
            ('NANUC', 'NANUC'),
            ('SÃO JOÃO DEL REI', 'SÃO JOÃO DEL REI'),
            ('BARBACENA', 'BARBACENA'),
            ('CONSELHEIRO LAFAIETE', 'CONSELHEIRO LAFAIETE'),
            ('OURO PRETO', 'OURO PRETO'),
            ('MARIANA', 'MARIANA')
        ]
        
        for key, microregion in microregion_mapping:
            if key in centro_str:
                return microregion
        
        return centro_str

    def _extrair_microregiao_from_rota(self, rota: str) -> str:
        """Extrai micro-região da rota (formato: ORIGEM-DESTINO)"""
        if pd.isna(rota):
            return "UNKNOWN"
        
        rota_str = str(rota).upper()
        
        # Se a rota contém "-", extrair a origem
        if "-" in rota_str:
            origem = rota_str.split("-")[0].strip()
        else:
            origem = rota_str
        
        # Mapeamento de micro-regiões - Minas Gerais
        microregion_mapping = [
            ('JOÃO MONLEVADE', 'JOÃO MONLEVADE'),
            ('USINA MONLEVADE', 'JOÃO MONLEVADE'),
            ('ITABIRA', 'ITABIRA'),
            ('BELO HORIZONTE', 'BELO HORIZONTE'),
            ('CONTAGEM', 'CONTAGEM'),
            ('SABARÁ', 'SABARÁ'),
            ('SANTA LUZIA', 'SANTA LUZIA'),
            ('NOVA LIMA', 'NOVA LIMA'),
            ('BRUMADINHO', 'BRUMADINHO'),
            ('IBIRITÉ', 'IBIRITÉ'),
            ('BETIM', 'BETIM'),
            ('LAGOA SANTA', 'LAGOA SANTA'),
            ('VESPASIANO', 'VESPASIANO'),
            ('RIBEIRÃO DAS NEVES', 'RIBEIRÃO DAS NEVES'),
            ('CAETÉ', 'CAETÉ'),
            ('SÃO JOSÉ DA LAPA', 'SÃO JOSÉ DA LAPA'),
            ('FLORESTAL', 'FLORESTAL'),
            ('JABOTICATUBAS', 'JABOTICATUBAS'),
            ('MATEUS LEME', 'MATEUS LEME'),
            ('IGARAPÉ', 'IGARAPÉ'),
            ('SÃO JOAQUIM DE BICAS', 'SÃO JOAQUIM DE BICAS'),
            ('SÃO JOSÉ DO GOIABAL', 'SÃO JOSÉ DO GOIABAL'),
            ('MARAVILHAS', 'MARAVILHAS'),
            ('ONÇA DE PITANGUI', 'ONÇA DE PITANGUI'),
            ('PARÁ DE MINAS', 'PARÁ DE MINAS'),
            ('PITANGUI', 'PITANGUI'),
            ('CONCEIÇÃO DO MATO DENTRO', 'CONCEIÇÃO DO MATO DENTRO'),
            ('SANTANA DO PARAÍSO', 'SANTANA DO PARAÍSO'),
            ('CORONEL FABRICIANO', 'CORONEL FABRICIANO'),
            ('IPATINGA', 'IPATINGA'),
            ('TIMÓTEO', 'TIMÓTEO'),
            ('CARATINGA', 'CARATINGA'),
            ('INHAPIM', 'INHAPIM'),
            ('GOVERNADOR VALADARES', 'GOVERNADOR VALADARES'),
            ('TEÓFILO OTONI', 'TEÓFILO OTONI'),
            ('NANUC', 'NANUC'),
            ('SÃO JOÃO DEL REI', 'SÃO JOÃO DEL REI'),
            ('BARBACENA', 'BARBACENA'),
            ('CONSELHEIRO LAFAIETE', 'CONSELHEIRO LAFAIETE'),
            ('OURO PRETO', 'OURO PRETO'),
            ('MARIANA', 'MARIANA')
        ]
        
        for key, microregion in microregion_mapping:
            if key in origem:
                return microregion
        
        return origem
    
    def analisar_tendencia_microregiao(self, microregiao: str, 
                                     periodo_meses: int = 12) -> Dict:
        """Analisa tendência de preços para uma micro-região específica"""
        if microregiao not in self.microregioes_disponiveis:
            return {"erro": f"Micro-região '{microregiao}' não encontrada"}
        
        # Filtrar dados da micro-região
        df_micro = self.df[self.df['microregiao_origem'] == microregiao].copy()
        
        if len(df_micro) == 0:
            return {"erro": f"Nenhum dado encontrado para micro-região '{microregiao}'"}
        
        # Calcular estatísticas por período
        stats_mensal = df_micro.groupby('ano_mes').agg({
            'preco_ton_km': ['mean', 'median', 'std', 'count'],
            'volume_ton': 'sum',
            'distancia_km': 'mean'
        }).round(4)
        
        # Flatten column names
        stats_mensal.columns = ['_'.join(col).strip() for col in stats_mensal.columns]
        
        # Reset index para facilitar análise
        stats_mensal = stats_mensal.reset_index()
        stats_mensal['ano_mes_str'] = stats_mensal['ano_mes'].astype(str)
        
        # Ordenar por período
        stats_mensal = stats_mensal.sort_values('ano_mes')
        
        # Calcular tendência linear
        if len(stats_mensal) >= 2:
            x = np.arange(len(stats_mensal))
            y = stats_mensal['preco_ton_km_mean'].values
            
            # Regressão linear
            slope, intercept = np.polyfit(x, y, 1)
            
            # Calcular R²
            y_pred = slope * x + intercept
            r_squared = 1 - np.sum((y - y_pred) ** 2) / np.sum((y - np.mean(y)) ** 2)
            
            # Determinar direção da tendência
            if slope > 0.001:
                direcao = "AUMENTANDO"
                tendencia_forca = "FORTE" if abs(slope) > 0.01 else "MODERADA"
            elif slope < -0.001:
                direcao = "DIMINUINDO"
                tendencia_forca = "FORTE" if abs(slope) > 0.01 else "MODERADA"
            else:
                direcao = "ESTÁVEL"
                tendencia_forca = "ESTÁVEL"
            
            # Calcular preço recomendado
            preco_atual = stats_mensal['preco_ton_km_mean'].iloc[-1]
            preco_medio_geral = df_micro['preco_ton_km'].mean()
            
            if direcao == "AUMENTANDO":
                if preco_atual < preco_medio_geral:
                    preco_recomendado = preco_atual * 1.05  # Aumentar 5%
                    acao = "AUMENTAR"
                else:
                    preco_recomendado = preco_atual
                    acao = "MANTER"
            elif direcao == "DIMINUINDO":
                if preco_atual > preco_medio_geral:
                    preco_recomendado = preco_atual * 0.95  # Reduzir 5%
                    acao = "REDUZIR"
                else:
                    preco_recomendado = preco_atual
                    acao = "MANTER"
            else:
                preco_recomendado = preco_atual
                acao = "MANTER"
            
        else:
            slope = 0
            r_squared = 0
            direcao = "INSUFICIENTE_DADOS"
            tendencia_forca = "N/A"
            preco_recomendado = df_micro['preco_ton_km'].mean()
            acao = "ANALISAR_MAIS_DADOS"
        
        # Últimos faturamentos
        ultimos_faturamentos = df_micro.nlargest(5, 'data_faturamento')[
            ['data_faturamento', 'preco_ton_km', 'volume_ton', 'distancia_km', 'rota_mesoregiao']
        ].copy()
        
        # Formatar datas
        ultimos_faturamentos['data_faturamento'] = ultimos_faturamentos['data_faturamento'].dt.strftime('%Y-%m-%d')
        
        # Estatísticas gerais
        stats_geral = {
            'total_rotas': len(df_micro),
            'periodo_analise': f"{stats_mensal['ano_mes'].min()} a {stats_mensal['ano_mes'].max()}",
            'preco_medio_geral': round(df_micro['preco_ton_km'].mean(), 4),
            'preco_mediano_geral': round(df_micro['preco_ton_km'].median(), 4),
            'desvio_padrao_geral': round(df_micro['preco_ton_km'].std(), 4),
            'volume_total_ton': round(df_micro['volume_ton'].sum(), 2),
            'distancia_media_km': round(df_micro['distancia_km'].mean(), 2)
        }
        
        return {
            'microregiao': microregiao,
            'tendencia': {
                'direcao': direcao,
                'forca': tendencia_forca,
                'inclinacao': round(slope, 6),
                'r_quadrado': round(r_squared, 4),
                'confiabilidade': "ALTA" if r_squared > 0.7 else "MÉDIA" if r_squared > 0.4 else "BAIXA"
            },
            'precos': {
                'atual': round(stats_mensal['preco_ton_km_mean'].iloc[-1], 4),
                'recomendado': round(preco_recomendado, 4),
                'acao': acao,
                'diferenca_percentual': round(((preco_recomendado - stats_mensal['preco_ton_km_mean'].iloc[-1]) / stats_mensal['preco_ton_km_mean'].iloc[-1]) * 100, 2)
            },
            'estatisticas_gerais': stats_geral,
            'evolucao_mensal': stats_mensal.to_dict('records'),
            'ultimos_faturamentos': ultimos_faturamentos.to_dict('records')
        }
    
    def gerar_relatorio_completo(self, output_path: str = "outputs"):
        """Gera relatório completo de todas as micro-regiões"""
        os.makedirs(output_path, exist_ok=True)
        
        print(f"\n📊 Analisando tendências para {len(self.microregioes_disponiveis)} micro-regiões...")
        
        resultados = {}
        resumo = []
        
        for i, microregiao in enumerate(self.microregioes_disponiveis, 1):
            print(f"  [{i:2d}/{len(self.microregioes_disponiveis)}] Analisando: {microregiao}")
            resultado = self.analisar_tendencia_microregiao(microregiao)
            
            if 'erro' not in resultado:
                resultados[microregiao] = resultado
                
                # Adicionar ao resumo
                resumo.append({
                    'microregiao': microregiao,
                    'tendencia': resultado['tendencia']['direcao'],
                    'forca': resultado['tendencia']['forca'],
                    'preco_atual': resultado['precos']['atual'],
                    'preco_recomendado': resultado['precos']['recomendado'],
                    'acao': resultado['precos']['acao'],
                    'confiabilidade': resultado['tendencia']['confiabilidade'],
                    'total_rotas': resultado['estatisticas_gerais']['total_rotas'],
                    'diferenca_percentual': resultado['precos']['diferenca_percentual']
                })
        
        # Criar DataFrame de resumo
        df_resumo = pd.DataFrame(resumo)
        
        # Ordenar por diferença percentual (maior impacto primeiro)
        df_resumo = df_resumo.sort_values('diferenca_percentual', ascending=False)
        
        # Salvar resultados
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Salvar resumo em Excel
        excel_path = os.path.join(output_path, f"tendencia_tku_resumo_{timestamp}.xlsx")
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='w') as writer:
            df_resumo.to_excel(writer, sheet_name='Resumo', index=False)
            
            # Adicionar detalhes por micro-região
            for microregiao, resultado in resultados.items():
                if 'erro' not in resultado:
                    # Dados mensais
                    df_mensal = pd.DataFrame(resultado['evolucao_mensal'])
                    df_mensal.to_excel(writer, sheet_name=f'{microregiao[:20]}_Mensal', index=False)
                    
                    # Últimos faturamentos
                    df_faturamentos = pd.DataFrame(resultado['ultimos_faturamentos'])
                    df_faturamentos.to_excel(writer, sheet_name=f'{microregiao[:20]}_Faturamentos', index=False)
        
        # Salvar JSON completo
        json_path = os.path.join(output_path, f"tendencia_tku_completo_{timestamp}.json")
        import json
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(resultados, f, ensure_ascii=False, indent=2, default=str)
        
        # Mostrar resumo dos resultados
        print(f"\n📋 RESUMO DAS TENDÊNCIAS:")
        print("=" * 80)
        print(f"{'Micro-região':<25} {'Tendência':<12} {'Preço Atual':<12} {'Recomendado':<12} {'Ação':<10} {'Impacto':<8}")
        print("=" * 80)
        
        for _, row in df_resumo.iterrows():
            impacto_emoji = "🔴" if abs(row['diferenca_percentual']) > 5 else "🟡" if abs(row['diferenca_percentual']) > 2 else "🟢"
            print(f"{row['microregiao']:<25} {row['tendencia']:<12} R${row['preco_atual']:<11} R${row['preco_recomendado']:<11} {row['acao']:<10} {impacto_emoji}")
        
        print("=" * 80)
        print(f"🔴 Impacto ALTO (>5%) | 🟡 Impacto MÉDIO (2-5%) | 🟢 Impacto BAIXO (<2%)")
        
        logger.info(f"Relatório salvo em: {excel_path}")
        logger.info(f"JSON completo salvo em: {json_path}")
        
        return {
            'excel_path': excel_path,
            'json_path': json_path,
            'resumo': df_resumo,
            'resultados': resultados
        }

    def gerar_resumo_oportunidades(self, periodo_meses: int = 12) -> pd.DataFrame:
        """
        Gera resumo de oportunidades de redução de custos por rota
        Baseado no dashboard do operador com corte temporal configurável
        
        Args:
            periodo_meses: Período de análise em meses (3, 6 ou 12)
        """
        logger.info(f"Gerando resumo de oportunidades para {periodo_meses} meses...")
        
        # Calcular data de corte
        data_atual = self.df['data_faturamento'].max()
        data_corte = data_atual - pd.DateOffset(months=periodo_meses)
        
        # Filtrar dados pelo período
        df_periodo = self.df[self.df['data_faturamento'] >= data_corte].copy()
        
        if len(df_periodo) == 0:
            logger.warning(f"Nenhum dado encontrado para o período de {periodo_meses} meses")
            return pd.DataFrame()
        
        logger.info(f"Dados filtrados para {periodo_meses} meses: {len(df_periodo)} linhas")
        
        # Agrupar por rota (apenas destino, como no script Spark)
        df_rotas = df_periodo.groupby(['rota_microregiao']).agg({
            'volume_ton': 'sum',
            'distancia_km': 'mean',
            'preco_ton_km': 'mean',
            'frete_brl': 'sum'
        }).reset_index()
        
        # Adicionar microregiao_origem baseada no centro_origem (primeiro valor encontrado)
        df_rotas['microregiao_origem'] = df_rotas['rota_microregiao'].apply(
            lambda x: self._extrair_microregiao_from_rota(x)
        )
        
        # Calcular custo superior (TKU)
        df_rotas['custo_sup_tku'] = df_rotas['preco_ton_km']
        
        # Calcular preço médio por micro-região de origem para o período
        # Usar a mesma lógica de extração de micro-região
        df_periodo['microregiao_origem_temp'] = df_periodo['centro_origem'].apply(self._extrair_microregiao)
        preco_medio_microregiao = df_periodo.groupby('microregiao_origem_temp')['preco_ton_km'].mean().reset_index()
        preco_medio_microregiao = preco_medio_microregiao.rename(columns={
            'microregiao_origem_temp': 'microregiao_origem',
            'preco_ton_km': 'preco_medio_microregiao'
        })
        
        # Mesclar com preços médios
        df_rotas = df_rotas.merge(
            preco_medio_microregiao, 
            on='microregiao_origem', 
            how='left'
        )
        
        # Calcular oportunidade (diferença entre custo atual e referência)
        df_rotas['oportunidade_brl_ton_km'] = df_rotas['custo_sup_tku'] - df_rotas['preco_medio_microregiao']
        
        # Determinar ação baseada na oportunidade
        def determinar_acao(oportunidade):
            if oportunidade > 0.01:  # Mais de 1 centavo de diferença
                return "Redução"
            elif oportunidade < -0.01:
                return "Aumento"
            else:
                return "Manter"
        
        df_rotas['acao'] = df_rotas['oportunidade_brl_ton_km'].apply(determinar_acao)
        
        # Calcular impacto estratégico
        def calcular_impacto(oportunidade, volume, distancia):
            if pd.isna(oportunidade) or pd.isna(volume) or pd.isna(distancia):
                return "BAIXO"
            
            impacto_score = abs(oportunidade) * volume * distancia / 1000000
            
            if impacto_score > 1000:
                return "ALTO"
            elif impacto_score > 100:
                return "MÉDIO"
            else:
                return "BAIXO"
        
        df_rotas['impacto_estrategico'] = df_rotas.apply(
            lambda row: calcular_impacto(
                row['oportunidade_brl_ton_km'], 
                row['volume_ton'], 
                row['distancia_km']
            ), axis=1
        )
        
        # Adicionar informações temporais
        df_rotas['periodo_analise'] = f"{periodo_meses} meses"
        df_rotas['data_inicio'] = data_corte.strftime('%Y-%m-%d')
        df_rotas['data_fim'] = data_atual.strftime('%Y-%m-%d')
        
        # Ordenar por volume (maior impacto primeiro)
        df_rotas = df_rotas.sort_values('volume_ton', ascending=False)
        
        # Selecionar colunas para o dashboard
        df_dashboard = df_rotas[[
            'rota_microregiao',
            'microregiao_origem',
            'volume_ton',
            'distancia_km',
            'custo_sup_tku',
            'preco_medio_microregiao',
            'oportunidade_brl_ton_km',
            'acao',
            'impacto_estrategico',
            'periodo_analise',
            'data_inicio',
            'data_fim'
        ]].copy()
        
        # Renomear colunas para o dashboard
        df_dashboard = df_dashboard.rename(columns={
            'rota_microregiao': 'Rota',
            'microregiao_origem': 'Micro-Região Origem',
            'volume_ton': 'Volume (TON)',
            'distancia_km': 'Distância (KM)',
            'custo_sup_tku': 'Custo Sup (TKU)',
            'preco_medio_microregiao': '04.01 - Média MicroRegião - Preço SUP (BRL/TON/KM)',
            'oportunidade_brl_ton_km': 'Oport. (BRL/TON/KM)',
            'acao': 'Ação',
            'impacto_estrategico': 'Impacto Estratégico',
            'periodo_analise': 'Período Análise',
            'data_inicio': 'Data Início',
            'data_fim': 'Data Fim'
        })
        
        logger.info(f"Resumo de oportunidades gerado para {periodo_meses} meses: {len(df_dashboard)} rotas")
        return df_dashboard
    
    def gerar_resumo_oportunidades_multiplos_periodos(self) -> Dict[str, pd.DataFrame]:
        """
        Gera resumo de oportunidades para múltiplos períodos (3, 6 e 12 meses)
        """
        logger.info("Gerando resumos de oportunidades para múltiplos períodos...")
        
        periodos = [3, 6, 12]
        resultados = {}
        
        for periodo in periodos:
            df_periodo = self.gerar_resumo_oportunidades(periodo)
            if len(df_periodo) > 0:
                resultados[f"{periodo}_meses"] = df_periodo
                logger.info(f"Período {periodo} meses: {len(df_periodo)} rotas")
            else:
                logger.warning(f"Período {periodo} meses: sem dados")
        
        return resultados
    
    def salvar_dashboard_oportunidades_multiplos_periodos(self, output_path: str = "outputs/dashboard_oportunidades_multiplos_periodos.xlsx"):
        """
        Salva o dashboard de oportunidades para múltiplos períodos em Excel
        """
        try:
            # Criar diretório se não existir
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # Gerar resumos para múltiplos períodos
            resultados = self.gerar_resumo_oportunidades_multiplos_periodos()
            
            if not resultados:
                logger.error("Nenhum resultado gerado para múltiplos períodos")
                return None
            
            # Salvar em Excel com múltiplas abas
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for periodo, df_periodo in resultados.items():
                    sheet_name = f"Oportunidades_{periodo}"
                    df_periodo.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Aplicar formatação
                    workbook = writer.book
                    worksheet = writer.sheets[sheet_name]
                    
                    # Formatar cabeçalho
                    from openpyxl.styles import Font, PatternFill, Alignment
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_font = Font(color="FFFFFF", bold=True)
                    
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center")
                    
                    # Formatar coluna de oportunidade (vermelho para valores positivos)
                    for row in range(2, len(df_periodo) + 2):
                        oportunidade_cell = worksheet.cell(row=row, column=6)  # Coluna Oportunidade
                        if oportunidade_cell.value and oportunidade_cell.value > 0:
                            oportunidade_cell.font = Font(color="FF0000", bold=True)
                    
                    # Ajustar largura das colunas
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Criar aba de resumo comparativo
                df_resumo_comparativo = self._criar_resumo_comparativo_periodos(resultados)
                if len(df_resumo_comparativo) > 0:
                    df_resumo_comparativo.to_excel(writer, sheet_name='Resumo_Comparativo', index=False)
                    
                    # Formatar aba de resumo
                    worksheet_resumo = writer.sheets['Resumo_Comparativo']
                    for cell in worksheet_resumo[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center")
            
            logger.info(f"Dashboard multi-período salvo em: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Erro ao salvar dashboard multi-período: {str(e)}")
            return None
    
    def _criar_resumo_comparativo_periodos(self, resultados: Dict[str, pd.DataFrame]) -> pd.DataFrame:
        """
        Cria resumo comparativo entre os diferentes períodos
        """
        resumos = []
        
        for periodo, df_periodo in resultados.items():
            if len(df_periodo) == 0:
                continue
                
            # Estatísticas por período
            total_rotas = len(df_periodo)
            rotas_reducao = len(df_periodo[df_periodo['Ação'] == 'Redução'])
            rotas_aumento = len(df_periodo[df_periodo['Ação'] == 'Aumento'])
            rotas_manter = len(df_periodo[df_periodo['Ação'] == 'Manter'])
            
            # Volume total e oportunidades
            volume_total = df_periodo['Volume (TON)'].sum()
            oportunidades_reducao = df_periodo[df_periodo['Ação'] == 'Redução']['Oport. (BRL/TON/KM)'].sum()
            
            resumos.append({
                'Período': periodo,
                'Total Rotas': total_rotas,
                'Rotas Redução': rotas_reducao,
                'Rotas Aumento': rotas_aumento,
                'Rotas Manter': rotas_manter,
                'Volume Total (TON)': volume_total,
                'Oportunidades Redução (BRL/TON/KM)': oportunidades_reducao,
                'Taxa Redução (%)': (rotas_reducao / total_rotas * 100) if total_rotas > 0 else 0
            })
        
        return pd.DataFrame(resumos)

def main():
    """Função principal"""
    print("🚀 Iniciando Análise de Tendência de Preços TKU")
    print("=" * 60)
    
    # Inicializar analisador
    analisador = AnalisadorTendenciaTKU()
    
    # Carregar dados
    if not analisador.carregar_dados():
        print("❌ Erro ao carregar dados. Verifique o arquivo sample_data.xlsx")
        return
    
    # Gerar relatório completo automaticamente
    print(f"\n🎯 Gerando relatório completo de tendências...")
    resultado = analisador.gerar_relatorio_completo()
    
    # Gerar dashboard de oportunidades
    print(f"\n📊 Gerando dashboard de oportunidades...")
    dashboard_path = analisador.salvar_dashboard_oportunidades_multiplos_periodos()
    
    if dashboard_path:
        print(f"✅ Dashboard de oportunidades salvo em: {dashboard_path}")
        
        # Mostrar resumo das oportunidades
        df_oportunidades = analisador.gerar_resumo_oportunidades_multiplos_periodos()
        
        print(f"\n🎯 RESUMO DE OPORTUNIDADES:")
        print(f"   • Total de rotas analisadas: {len(df_oportunidades)}")
        
        # Contar ações
        acoes = df_oportunidades['Ação'].value_counts()
        for acao, count in acoes.items():
            print(f"   • {acao}: {count} rotas")
        
        # Top 5 maiores oportunidades de redução
        df_reducao = df_oportunidades[df_oportunidades['Ação'] == 'Redução'].sort_values('Oport. (BRL/TON/KM)', ascending=False)
        if len(df_reducao) > 0:
            print(f"\n🔥 TOP 5 OPORTUNIDADES DE REDUÇÃO:")
            for i, (_, row) in enumerate(df_reducao.head(5).iterrows(), 1):
                print(f"   {i}. {row['Centro Origem']}: R$ {row['Oport. (BRL/TON/KM)']:.4f} "
                      f"(Volume: {row['Volume (TON)']:,.1f} ton)")
        
        # Distribuição por impacto
        impacto_dist = df_oportunidades['Impacto Estratégico'].value_counts()
        print(f"\n📈 DISTRIBUIÇÃO POR IMPACTO:")
        for impacto, count in impacto_dist.items():
            print(f"   • {impacto}: {count} rotas")
    
    print(f"\n✅ Análise concluída!")
    print(f"📁 Relatório Excel: {resultado['excel_path']}")
    print(f"📄 Dados JSON: {resultado['json_path']}")
    
    # Estatísticas finais
    df_resumo = resultado['resumo']
    total_aumentar = len(df_resumo[df_resumo['acao'] == 'AUMENTAR'])
    total_reduzir = len(df_resumo[df_resumo['acao'] == 'REDUZIR'])
    total_manter = len(df_resumo[df_resumo['acao'] == 'MANTER'])
    
    print(f"\n📊 RESUMO FINAL:")
    print(f"   • Micro-regiões analisadas: {len(df_resumo)}")
    print(f"   • Recomendações AUMENTAR: {total_aumentar}")
    print(f"   • Recomendações REDUZIR: {total_reduzir}")
    print(f"   • Recomendações MANTER: {total_manter}")
    
    # Top 5 maiores oportunidades
    print(f"\n🔥 TOP 5 MAIORES OPORTUNIDADES:")
    for i, (_, row) in enumerate(df_resumo.head(5).iterrows(), 1):
        print(f"   {i}. {row['microregiao']}: {row['acao']} {row['diferenca_percentual']:+.1f}% "
              f"(R$ {row['preco_atual']:.4f} → R$ {row['preco_recomendado']:.4f})")

if __name__ == "__main__":
    main()
